# -*- coding: utf-8 -*-

import requests
import os
from datetime import datetime
from random import random, choice
from hashlib import md5
import json
import openpyxl
import logging
from lxml import etree
import pandas as pd
import numpy as np


# 第一步，创建一个logger
logger = logging.getLogger()
logger.setLevel(logging.INFO)  # Log等级总开关

# 第二步，创建一个handler，用于写入日志文件
logfile = './log2.txt'
fh = logging.FileHandler(logfile, mode='a')
fh.setLevel(logging.DEBUG)  # 用于写到file的等级开关

# 第三步，再创建一个handler,用于输出到控制台
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)  # 输出到console的log等级的开关

# 第四步，定义handler的输出格式
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
fh.setFormatter(formatter)
ch.setFormatter(formatter)

# 第五步，将logger添加到handler里面
logger.addHandler(fh)
logger.addHandler(ch)


conf_newrank = {
    'username': '*****',
    'password': '*****'
}

# 读取excel文件
row_excel = 0
count_0 = []
class NewRankApi(object):
    __headers = {
        'origin': 'https://www.newrank.cn',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.119 Safari/537.36',
        'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'authority': 'www.newrank.cn',
        'accept': 'application/json, text/javascript, */*; q=0.01',
        'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8',
        'x-requested-with': 'XMLHttpRequest'
    }

    def __init__(self):
        self.__data_file = os.path.join('.', 'data.json')
        if not os.path.exists(self.__data_file):
            newrank_user = self._login()
        else:
            with open(self.__data_file, 'r') as f:
                data = f.read()
                newrank_user = json.loads(data) if data else None
                self._build_cookies(newrank_user)
            # 测试用户登录态是否失效
            if not self._check_online():
                newrank_user = self._login()
        self._build_cookies(newrank_user)
    @staticmethod
    def read_excel(rows, col):  # 22,3
        total_list = []
        wb = openpyxl.load_workbook("./3月12日20个发布链接抓阅读量.xlsx")
        sheet = wb['Sheet1']
        for r in range(2, rows):
            row_list = []  # 每一行建立一个list
            for c in range(1, col + 1):
                v = sheet.cell(r, c).value
                row_list.append(v)
            total_list.append(row_list)
        wb.close()
        print(total_list)
        return total_list

    # 获取excel中的文章链接
    @staticmethod
    def get_rank(rows,col):  # 22，3
        total_list = NewRankApi.read_excel(rows,col)  #22，3
        # total_list = NewRankApi.read_excel()

        rank = total_list[row_excel][2]
        return rank

    # 从excel中获取文章标题和公众名称
    @staticmethod
    def analysis_link(rows,col): #22,3
        link = NewRankApi.get_rank(rows,col) #22,3
        response = requests.get(link)
        html_str = response.text
        html = etree.HTML(html_str)
        list_v = []
        theme = html.xpath("//div/h2[@id='activity-name']/text()")[0].replace("\n", "").strip()
        list_v.append(theme)
        gzh_name = html.xpath("//div/span[@id='profileBt']/a[@id='js_name']/text()")[0].replace("\n", "").strip()
        list_v.append(gzh_name)
        return list_v

    def _build_cookies(self, newrank_user):
        '''构造请求 cookie'''
        if not newrank_user:
            raise Exception('构造cookie时，传递了空的用户信息')
        self.__cookies = {
            'rmbuser': 'true',
            'name': newrank_user.get('phone'),
            'token': newrank_user.get('token'),
            'openid': newrank_user.get('wxopenid'),
            'tt_token': 'true',
            'useLoginAccount': 'true'
        }
        return self.__cookies


    def _get_url(self, request_uri):
        return 'https://www.newrank.cn' + request_uri

    def _rebuild_data(self, request_uri, data):
        '''处理请求参数'''
        # 排序
        keys = list(data.keys())
        keys.sort()
        data['nonce'] = ''.join(choice('0123456789abcdef') for x in range(0, 9))
        keys.append('nonce')
        # 转换 bool 类型值为 str，并计算xyz
        l = []
        for k in keys:
            v = data.get(k)
            if isinstance(v, bool):
                v = str(v)
                v = v[0].lower() + v[1:]
                data[k] = v
            l.append(k + '=' + v)
        data['xyz'] = md5((request_uri + '?AppKey=joker&' + '&'.join(l)).encode('utf-8')).hexdigest()
        return data

    def _check_online(self, tries=1):
        '''判断用户是否处于登录态'''
        # 构造请求参数
        request_uri = '/xdnphb/common/account/get'
        headers = {
            'referer': self._get_url('/'),
            **self.__headers
        }
        data = {}
        self._rebuild_data(request_uri, data)
        # 发起请求
        resp = requests.post(self._get_url(request_uri), headers=headers, data=data, cookies=self.__cookies, timeout=30)
        try:
            error_str = '校验用户信息接口响应错误: ' + str(resp.status_code) + ' ' + resp.text
            if resp.status_code != 200:
                raise Exception(error_str)
            resp_data = resp.json()
            if resp_data.get('success') != True:
                raise Exception(error_str)
        except Exception:
            if tries >= 3:
                raise Exception('校验用户信息失败')
            else:
                tries += 1
                return self._check_online(tries)
        return isinstance(resp_data.get('value'), dict)

    def _login(self):
        '''
        登录接口
        username - 登录新榜的用户名
        password - 登录新榜的密码
        '''
        # 构造请求参数
        request_uri = '/xdnphb/login/new/usernameLogin'
        headers = {
            'referer': self._get_url('/public/login/login.html?back=https%3A//www.newrank.cn/'),
            **self.__headers
        }
        flag = str(int(datetime.now().timestamp() * 1000)) + str(random())
        hash_password = md5(
            (md5(conf_newrank['password'].encode('utf-8')).hexdigest() + 'daddy').encode('utf-8')).hexdigest()
        data = {
            'flag': flag,
            'identifyCode': '',
            'password': hash_password,
            'username': conf_newrank['username'],
        }
        self._rebuild_data(request_uri, data)
        # 发起请求
        resp = requests.post(self._get_url(request_uri), headers=headers, data=data, timeout=30)
        error_str = '登录接口响应错误: ' + str(resp.status_code) + ' ' + resp.text
        if resp.status_code != 200:
            raise Exception(error_str)
        try:
            resp_data = resp.json()
        except Exception:
            raise Exception(error_str)
        if resp_data.get('success') != True:
            raise Exception(error_str)
        newrank_user = resp_data.get('value')
        with open(self.__data_file, 'w') as f:
            f.write(json.dumps(newrank_user))
        logger.info('登录并获取用户信息，写入数据文件')
        return newrank_user

    def query(self, kw,tries=1):
        '''
        获取新榜分配给公众号的 uuid，提供给 download 接口使用

        kw - 公众号名称
        '''

        # 构造请求参数
        request_uri = '/xdnphb/data/weixinuser/searchWeixinDataByCondition'
        headers = {
            'referer': self._get_url('/public/info/search.html?value=' + kw + '&isBind=false'),
            **self.__headers
        }
        data = {
            'filter': '',
            'hasDeal': False,
            'keyName': kw,
            'order': 'relation',
        }
        datas = self._rebuild_data(request_uri, data)
        urls = self._get_url(request_uri)
        # 发起请求

        resp = requests.post(url=urls, headers=headers, data=datas, cookies=self.__cookies, timeout=30)
        error_str = '公众号[' + kw + ']查询uuid接口响应错误: ' + str(resp.status_code) + ' ' + resp.text

        try:
            if resp.status_code != 200:
                raise Exception(error_str)
            resp_data = resp.json()
            account = resp_data["value"]['result'][0]['account']
        except Exception:
            if tries >=3:
                raise Exception('微信号抓取文章失败')
            else:
                tries += 1
                return self.query(kw, tries)
        logger.info("第二步：query(kw, tries=1),获得该公众号名对应的微信名称：{}".format(account))
        return account


    def get_articles(self,rows,col): #22,3
        kword = self.analysis_link(rows,col)[1]
        account1 = self.query(kword,tries=1)
        '''
        下载数据接口
        account - 公众号的微信号
        '''
        # 构造请求参数
        request_uri = '/xdnphb/detail/v1/rank/article/lists'
        headers = {
            'referer':self._get_url('/new/detial?account={}'.format(account1)),
            'origin': 'https://www.newrank.cn',
            'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.119 Safari/537.36',
            'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'authority': 'www.newrank.cn',
            'accept': 'application/json, text/javascript, */*; q=0.01',
            'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8',
            'x-requested-with': 'XMLHttpRequest'
        }
        data = {
            'account':account1
        }
        datas = self._rebuild_data(request_uri, data)
        url = self._get_url(request_uri)
        # 发起请求，处理响应
        resp = requests.post(url=url, headers=headers, data=datas, cookies=self.__cookies, timeout=30)
        resp_data = resp.json()
        articles_list = resp_data['value']['articles']
        logger.info("第三步：获得（{}）公众下最近7天的文章列表，————{}——".format(kword,articles_list))
        return articles_list

    # def get_read_count(self):
    #     global count_0
    #     theme = NewRankApi.analysis_link()[0]
    #     # logger.info("第一步 analysis_link()：根据excel链接获得文章标题和公众号名称：{}".format(theme))
    #
    #
    #     # kword = NewRankApi.analysis_link()[1]
    #     articles = NewRankApi.get_articles(self)
    #     print(articles)
    #     print(type(articles))
    #     for out in articles:
    #         for article in out:
    #             print(article['title'])
    #             # spider_theme = article['title'].replace("\n", "").strip()
    #             spider_theme = article['title']
    #             # print("公众号：{}".format(theme))
    #             # print(spider_theme)
    #             if theme==spider_theme:
    #                 logger.info("第四步：对比文章标题：爬取的文章--{} --- 原有文章{}".format(spider_theme,self.analysis_link()[0]))
    #                 clicksCounts = article['clicksCount']
    #
    #                 print(clicksCounts)
    #                 count_0.append(clicksCounts)
    #                 return clicksCounts
    #             else:
    #                 logger.info("第四步：对比文章标题失败")
    #                 clicksCounts = "no"
    #                 count_0.append(clicksCounts)
    #                 print(count_0)
    #                 return clicksCounts
    #     return "ok"

    def get_read_count(self,rows,col): #22,3
        global count_0
        try:
            theme = NewRankApi.analysis_link(rows,col)[0]
            articles = NewRankApi.get_articles(self,rows,col)
            for article in articles:
                for art in article:
                    if art["title"] ==theme:
                        clicksCounts = art['clicksCount']
                        count_0.append(clicksCounts)
                        logger.info("获取浏览数成功")
                        print(count_0)
                        return clicksCounts
                    else:
                        clicksCounts = "no"
                        count_0.append(clicksCounts)
                        logger.info("获取浏览数失败")
                        print(count_0)
                        return clicksCounts
        except:
            count_0.append("链接打开无效")


    def run(self,rows,col):  #22,3
        global row_excel,count_0
        while row_excel <= rows-3:
            NewRankApi.get_read_count(self,rows,col)
            row_excel += 1
        return count_0

if __name__ == '__main__':
    data = NewRankApi().run(22,3)
    wb = openpyxl.load_workbook("./3月12日20个发布链接抓阅读量.xlsx")
    sheet = wb['Sheet1']
    sheet.cell(row=1, column=4, value="浏览量")
    for i, item in enumerate(data):
        sheet.cell(row=i + 2, column=4, value=item)
    wb.save("3月12日20个发布链接抓阅读量.xlsx")



